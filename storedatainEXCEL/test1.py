import wmi
import socket
import uuid
from datetime import datetime
import os
import re
import psutil
import xlsxwriter
import openpyxl


def get_boot_time():
    # Get system boot time
    boot_time_timestamp = psutil.boot_time()
    boot_time = datetime.fromtimestamp(boot_time_timestamp).strftime("%Y-%m-%d %H:%M:%S")
    return boot_time


def get_system_info():
    # Get system information
    system = wmi.WMI()
    serial_number = system.Win32_BIOS()[0].SerialNumber.strip()
    host_name = socket.gethostname()
    ip_address = socket.gethostbyname(host_name)
    mac_address = ':'.join(re.findall('..', '%012x' % uuid.getnode()))
    now = datetime.now()
    timestamp = now.strftime("%Y-%m-%d %H:%M:%S")
    return serial_number, host_name, ip_address, mac_address, timestamp


def write_data_to_file(file_path, data):
    if os.path.exists(file_path):
        # Check if temp.txt file exists and can be uploaded
        if os.path.exists("temp.txt"):
            with open("temp.txt", "r") as f:
                text = f.read()
                text = text.split('\n')
                # print(text)
                loop_completed = True
                for line in text:
                    # print(line,"LINE")
                    try:
                        # Append data to existing file
                        wb = openpyxl.load_workbook("system_info.xlsx")
                        ws = wb.active
                        ws.append(line.strip().split(","))
                        wb.save("system_info.xlsx")
                    except:
                        loop_completed = False
                        print("Temp data cannot be uploaded at this moment.")
                        break
                if loop_completed:
                    # do something if loop ran to completion
                    print("Temp data uploaded successfully!")

        # Append present data to existing file
        try:
            # Load the workbook
            row = data
            wb = openpyxl.load_workbook("system_info.xlsx")
            ws = wb.active
            ws.append(row)
            wb.save("system_info.xlsx")
            print(f"Data saved successfully: {row}")

            # Delete the temp file
            try:
                os.remove("temp.txt")
            except:
                print("temp not present")
        except:
            # Create new temp file and add data to it
            if os.path.exists("temp.txt"):
                tempData = f"\n{','.join(row)}"
            else:
                tempData = f"{','.join(row)}"
            with open("temp.txt", "a") as k:
                k.write(tempData)
                
            print("Data added to temp file for later upload.")
    else:
        # Create new file and add data
        workbook = xlsxwriter.Workbook(file_path)
        worksheet = workbook.add_worksheet()
        worksheet.write_row(0, 0, ['Serial Number', 'Host Name', 'IP Address', 'MAC Address', 'Time'])
        row = data
        worksheet.write_row(1, 0, row)
        workbook.close()
        print("Data saved successfully.")


def remove_empty_rows(file_path):
    try:
        # Open the workbook and select the worksheet
        workbook = xlsxwriter.Workbook(file_path)
        worksheet = workbook.add_worksheet()

        # Get the maximum number of rows in the worksheet
        max_row = worksheet.dim_rowmax

        # Iterate over the rows in the worksheet
        for row in range(max_row, -1, -1):
            # Check if all cells in the row are empty
            if not worksheet.row_values(row):
                # Delete the row
                worksheet.delete_row(row)

        # Save the workbook
        workbook.close()

    except:
        print("Can't remove empty rows.")


def main():
    boot_time = get_boot_time()
    system_info = get_system_info()
    data = list(system_info)
    data.insert(4, boot_time)
    file_path = "system_info.xlsx"
    write_data_to_file(file_path,data)
    # remove_empty_rows(file_path)

main()

