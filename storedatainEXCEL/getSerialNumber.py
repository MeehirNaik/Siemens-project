import wmi
import socket
import uuid
from datetime import datetime
import os
import re
import openpyxl
import psutil
from datetime import datetime

boot_time_timestamp = psutil.boot_time()
boot_time = datetime.fromtimestamp(boot_time_timestamp).strftime("%Y-%m-%d %H:%M:%S")
print(f"System boot time: {boot_time}")

# Get system information
system = wmi.WMI()
serial_number = system.Win32_BIOS()[0].SerialNumber.strip()
host_name = socket.gethostname()
ip_address = socket.gethostbyname(host_name)
print(ip_address)
mac_address = ':'.join(re.findall('..', '%012x' % uuid.getnode()))
print(mac_address)
mac_address1 = ':'.join(['{:02x}'.format((uuid.getnode() >> ele) & 0xff) for ele in range(0,8*6,8)][::-1])
print(mac_address1)

# get current date and time
now = datetime.now()

# convert to string format
timestamp = now.strftime("%Y-%m-%d %H:%M:%S")

# Check if system_info.xlsx file exists
if os.path.exists("system_info.xlsx"):
    wb = openpyxl.load_workbook("system_info.xlsx")
    ws = wb.active
    # Check if temp.txt file exists and can be uploaded
    if os.path.exists("temp.txt"):
        with open("temp.txt", "r") as f:
            text = f.read()
            print(text)
            for line in text.split('\n'):
                try:
                    # Append data to existing file
                    ws.append(line.strip().split(","))
                    wb.save("system_info.xlsx")
                    print("Temp data uploaded successfully!")
                except:
                    print("Temp data cannot be uploaded at this moment.")

    # Append present data to existing file
    try:
        row = [serial_number, host_name, ip_address, mac_address, boot_time, timestamp]
        ws.append(row)
        wb.save("system_info.xlsx")
        print(f"Data saved successfully: {row}")

        # Delete the temp file
        try:
            os.remove("temp.txt")
        except:
            print("temp not present")
    except:
        # Create new temp file and add data to it
        print("im in EXEPT")

        with open("temp.txt", "a") as k:
            k.write(f"\n{serial_number},{host_name},{ip_address},{mac_address},{boot_time},{timestamp}")
        print("Data added to temp file for later upload.")
else:
    # Create new file and add data
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Serial Number', 'Host Name', 'IP Address', 'MAC Address', 'Boot Time','Time'])
    ws.append([serial_number, host_name, ip_address, mac_address, boot_time,timestamp])
    wb.save("system_info.xlsx")
    print("Data saved successfully.")

try:
    # Load the Excel file
    workbook = openpyxl.load_workbook("system_info.xlsx")

    # Select the active worksheet
    worksheet = workbook.active

    # Iterate over the rows and delete any empty row
    for row in reversed(list(worksheet.rows)):
        if all(cell.value is None or cell.value == "" for cell in row):
            worksheet.delete_rows(row[0].row, amount=1)

    # Save the modified file
    workbook.save("system_info.xlsx")

except:
    print("can't remove spaces")