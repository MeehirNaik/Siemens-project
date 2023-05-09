import wmi
import socket
import uuid
from datetime import datetime
import os
import re
import psutil
import xlsxwriter
import openpyxl
import xml.etree.ElementTree as ET
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment

startRow = 2

def get_boot_time():
    # Get system boot time
    boot_time_timestamp = psutil.boot_time()
    boot_time = datetime.fromtimestamp(boot_time_timestamp).strftime("%Y-%m-%d %H:%M:%S")
    return boot_time


def get_system_info():
    # Get system information
    system = wmi.WMI()
    serial_number = system.Win32_BIOS()[0].SerialNumber.strip()
    host_name = socket.gethostname()
    ip_address = socket.gethostbyname(host_name)
    mac_address = ':'.join(re.findall('..', '%012x' % uuid.getnode()))
    now = datetime.now()
    timestamp = now.strftime("%Y-%m-%d %H:%M:%S")
    return serial_number, host_name, ip_address, mac_address, timestamp

def createNewTextFile(tempFile,row):
    # Create new temp file and add data to it
            if os.path.exists(tempFile):
                tempData = f"\n{','.join(row)}"
            else:
                tempData = f"{','.join(row)}"
            with open(tempFile, "a") as k:
                k.write(tempData)

def readXmlFile(xml_file_path):

    try:
        # Parse the XML file
        tree = ET.parse(xml_file_path)
        # Get the root element
        root = tree.getroot()
        # Find the 'automate' element using its tag
        automate_elem = root.find('AUTOMAT')
        # Get the data from the 'automate' element
        automate_data = automate_elem.text
        # Print the data
        print(automate_data)
    
    except Exception as e:
        print(e)

    return automate_data

def addColor(ws,startRow):
    # Define the fill colors
    fill1 = PatternFill(start_color='CBB279', end_color='CBB279', fill_type='solid')
    fill2 = PatternFill(start_color='E1D4BB', end_color='E1D4BB', fill_type='solid')
    # Create a new alignment style object with center alignment
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # Add alternating fill colors to each row
    for k in range(startRow,(ws.max_row) + 1):
        ws.row_dimensions[k].height = 30
        if k % 2 == 0:
            fill = fill1
        else:
            fill = fill2
        for cell in ws[k]:    
            cell.fill = fill
            cell.alignment = center_alignment
            
    return ws

def write_data_to_file(file_path, data):
    global startRow
    if os.path.exists(file_path + "system_info.xlsx"):
        # Check if temp.txt file exists and can be uploaded
        if os.path.exists("temp" + file_path[-6:-2] + ".txt"):
            with open("temp" + file_path[-6:-2] + ".txt", "r") as f:
                text = f.read()
                text = text.split('\n')
                loop_completed = True
                for line in text:
                    try:
                        # Append data to existing file
                        wb = openpyxl.load_workbook(file_path + "system_info.xlsx")
                        ws = wb.active
                        ws.append(line.strip().split(","))
                        addColor(ws,startRow)
                        temp = ws.max_row
                        wb.save(file_path + "system_info.xlsx")
                        startRow = temp
                    except Exception as e:
                        print(e)
                        loop_completed = False
                        print("Temp data cannot be uploaded at this moment.")
                        break
            if loop_completed:
                # do something if loop ran to completion
                print("Temp data uploaded successfully!")
                # Delete the temp file once 
                os.remove("temp" + file_path[-6:-2] + ".txt")

        # Append present data to existing file
        try:
            # Load the workbook
            row = data
            wb = openpyxl.load_workbook("system_info.xlsx")
            ws = wb.active
            ws.append(row)
            addColor(ws, startRow)
            temp = ws.max_row
            wb.save(file_path + "system_info.xlsx")
            startRow = temp
            print(f"Data saved successfully: {row}")

        except Exception as e:
            print(e)
            createNewTextFile("temp" + file_path[-6:-2] + ".txt",row)
                
            print("Data added to temp file for later upload.")
    else:
        if not os.path.exists(file_path) and file_path != '':
            os.mkdir(file_path)
        # Create new file and add data
        workbook = xlsxwriter.Workbook(file_path + "system_info.xlsx")
        worksheet = workbook.add_worksheet()
        # Set the width of columns A and B to 20 pixels
        worksheet.set_column('A:G', 30)
        worksheet.set_default_row(30)

        # Define a format for the header row
        header_format = workbook.add_format({
            'bold': True,
            'align': 'center',
            'font_color': 'white',
            'bg_color': '#537188',  # set the background color to a light pink
        })

        # Write the headers
        worksheet.write_row(0, 0, 
                            ['Tester Name',
                            'Serial Number',
                            'Host Name',
                            'IP Address', 
                            'MAC Address', 
                            'Boot Time', 
                            'Read Time'], header_format)
        row = data
        worksheet.write_row(1, 0, row,workbook.add_format({'bg_color': '#CBB279'}))
        # Freeze the first row
        worksheet.freeze_panes(1, 0)
        workbook.close()
        print("Data saved successfully.")


def remove_empty_rows(file_path):
    try:
        # Open the workbook and select the worksheet
        workbook = xlsxwriter.Workbook(file_path)
        worksheet = workbook.add_worksheet()

        # Get the maximum number of rows in the worksheet
        max_row = worksheet.dim_rowmax

        # Iterate over the rows in the worksheet
        for row in range(max_row, -1, -1):
            # Check if all cells in the row are empty
            if not worksheet.row_values(row):
                # Delete the row
                worksheet.delete_row(row)

        # Save the workbook
        workbook.close()

    except:
        print("Can't remove empty rows.")


def main():
    boot_time = get_boot_time()
    system_info = get_system_info()
    data = list(system_info)
    data.insert(4, boot_time)
    AUTOMAT_file_path = "wups\\MLFB\\AUTOMAT.xml"
    testerName = readXmlFile(AUTOMAT_file_path)
    data.insert(0, testerName)
    file_path = ""
    file_path_on_server = "server\\"
    write_data_to_file(file_path,data)
    write_data_to_file(file_path_on_server,data)
    # remove_empty_rows(file_path)

main()